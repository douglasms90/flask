from openpyxl import Workbook, load_workbook


class worksheet():
  def __init__(self):
    self.workbook = Workbook()
    self.worksheet = self.workbook.active

  def add_ob(self, ob, row):
    self.worksheet.cell(column=1, row=row).value = ob["column1"]
    self.worksheet.cell(column=2, row=row).value = ob["column2"] 
    self.worksheet.cell(column=3, row=row).value = ob["column3"]
    self.worksheet.cell(column=4, row=row).value = ob["column4"] 
    self.worksheet.cell(column=5, row=row).value = ob["column5"]
    self.worksheet.cell(column=6, row=row).value = ob["column6"]
    self.worksheet.cell(column=7, row=row).value = ob["column7"]
    self.worksheet.cell(column=8, row=row).value = ob["column8"]
    self.worksheet.cell(column=9, row=row).value = ob["column9"]
    self.worksheet.cell(column=10, row=row).value = ob["column10"]

  def add_into_sheet(self, obj_list):
    for i, ob in enumerate(obj_list):
      self.add_ob(ob, i + 2)
  
  def save(self, path_file):
    self.workbook.save(path_file)
