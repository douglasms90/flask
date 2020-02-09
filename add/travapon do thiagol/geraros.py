#!/usr/bin/env python
# -*- coding: utf-8 -*- 
import sys
from openpyxl import load_workbook
import psycopg2
import psycopg2.extensions
psycopg2.extensions.register_type(psycopg2.extensions.UNICODE)
psycopg2.extensions.register_type(psycopg2.extensions.UNICODEARRAY)
import re
import datetime
import os
import shutil

#Date Formation 
data = datetime.date.today()+ datetime.timedelta(days=1)
data = data.strftime('%d/%m/%Y')

#File name XLSX
filename = datetime.date.today()+datetime.timedelta(days=1)
filename = filename.strftime('%d-%m-%Y')

shutil.copy2('/var/www/Flask/modelo.xlsx','/var/www/Flask/modelo2.xlsx')
os.rename('/var/www/Flask/modelo2.xlsx','/var/www/Flask/'+filename+'.xlsx')

#XLSX Loading
wb = load_workbook('/var/www/Flask/'+filename+'.xlsx')
ws = wb.active

#Database Connection
def databaseConnection():
 try:
  conn = psycopg2.connect("dbname='mkData3.0' user='cliente_r' host='177.184.72.6' password='Cl13nt_R'")
 except:
  return "Impossible to connect to the database, check your code."
 
 cur = conn.cursor()
 conn.set_client_encoding('LATIN1')
 cur.execute("SELECT os.codos, def.descricao_defeito, c.cidade, p.nome_razaosocial FROM mk_os os INNER JOIN mk_pessoas p ON os.cliente=p.codpessoa LEFT JOIN mk_os_defeitos def ON os.defeito_associado=def.coddefeito INNER JOIN mk_cidades c ON os.cd_cidade=c.codcidade where data_abertura=current_date AND status='1' AND tipo_os in ('4','15') ORDER BY os.codos ASC")
 #cur.execute("SELECT os.codos, def.descricao_defeito, c.cidade, p.nome_razaosocial FROM mk_os os INNER JOIN mk_pessoas p ON os.cliente=p.codpessoa INNER JOIN mk_os_defeitos def ON os.defeito_associado=def.coddefeito INNER JOIN mk_cidades c ON os.cd_cidade=c.codcidade where status='1' AND tipo_os in ('4','15') ORDER BY os.codos ASC LIMIT 50")
 rows = cur.fetchall()
  
 return rows
 
#Row Count Database
def checkRaw(row):
 if len(row) > 40:
  shutil.copy2('/var/www/Flask/modelo.xlsx','/var/www/Flask/modelo2.xlsx')
  os.rename('/var/www/Flask/modelo2.xlsx','/var/www/Flask/'+filename+'-2.xlsx')
 else:
  pass


def dumpData(rows):
 obj_list = list()
 for row in rows:
  ob = dict() 
  ob['ID'] = row[0]

#DESABILITE ESSA LINHA QUANDO O REGEX APRESENTAR UM ERRO.
  #ob['DEFEITO'] = row[1]
  #ob['CLIENTE'] = row[3]
  
  #Regex Abbreviation 
  if row[1]:
   if re.match(r'^Intermit',row[1]):
    ob['DEFEITO'] = 'INTERMITÊNCIA'   
   if re.match(r'^Manuten',row[1]):
    ob['DEFEITO'] = 'REDE'  
   else:
    ob['DEFEITO'] = row[1].upper()
  else:
   ob['DEFEITO'] = row[1]

  ob['CLIENTE'] = row[3]

  #Regex "Barra do Pirai / Pirai"
  if re.match(r'^B', row[2]):
   ob["BP"] = "X"  
   ob["PI"] = ""
  else:  
   ob["PI"] = "X"
   ob["BP"] = "" 
  obj_list.append(ob)
 
 return obj_list

def add_ob(ob,row):
 ws.cell(column=3, row=row).value = ob['ID']
 ws.cell(column=4, row=row).value = ob['DEFEITO']
 ws.cell(column=6, row=row).value = ob['BP'] 
 ws.cell(column=7, row=row).value = ob['PI'] 
 ws.cell(column=8, row=row).value = ob['CLIENTE']

def add_into_sheet(list):
 for i, ob in enumerate(list): 	
  add_ob(ob,i+13)

def main():

 #Insert DATA
 ws.cell(row=4,column=6).value = data

 #DUMP DATA
 rows = databaseConnection()

 obj_list = dumpData(rows)

 add_into_sheet(obj_list)

 wb.save('/var/www/Flask/'+filename+'.xlsx')
 print "Lista criada com sucesso!"	

if __name__ == "__main__":
 main()


