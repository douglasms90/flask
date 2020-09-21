from openpyxl import load_workbook
from datetime import date, timedelta
from psycopg2 import connect
from os import makedirs, path
import psycopg2.extensions
psycopg2.extensions.register_type(psycopg2.extensions.UNICODE)
psycopg2.extensions.register_type(psycopg2.extensions.UNICODEARRAY)
from re import match
import webbrowser


class databaseConnection:
  _db = None    
  def __init__(self, conn):
    self._db = psycopg2.connect(conn)

  def Consult(self, select):
    database = None 
    try:
      cur = self._db.cursor()
      cur.execute(select)
      database = cur.fetchall()
      return database
    except:
      return "Impossible to connect to the database, check your code."
  
  def close(self):
     self._db.close()

class workSheet():

  def __init__(self, path_xlsx):
    self.workbook = load_workbook(path_xlsx)
    self.worksheet = self.workbook.active

  def dumpData(self, database, tomorrow):
    obj_list = list()
    count = 0
    for data in database:
      ob = dict()
      count += 1
      ob["column1"] = count # Quantos
      ob["column2"] = data[0] # Id
      ob["column3"] = data[1] # Defeito
      ob["column4"] = data[2] # Tipo
      if data[3]:
        ob["column5"] = data[3].strftime("%d/%m/%Y") # Gerada
      else:
        ob["column5"] = "X"
      ob["column6"] = data[4] # Cliente
      ob["column7"] = ''
      ob["column8"] = tomorrow # Distribuida
      ob["column9"] = data[5] # Cidade
      ob["column10"] = data[6].title() # Bairro
      ob["column12"] = data[7].title() # Logradouro
      obj_list.append(ob)
    return obj_list

  def add_ob(self, ob, row):
    self.worksheet.cell(column=1, row=row).value = ob["column1"] # Quantidade
    self.worksheet.cell(column=2, row=row).value = ob["column2"] # Id
    self.worksheet.cell(column=3, row=row).value = ob["column3"] # Defeito
    self.worksheet.cell(column=4, row=row).value = ob["column4"] # Tipo
    self.worksheet.cell(column=5, row=row).value = ob["column5"] # Gerada
    self.worksheet.cell(column=6, row=row).value = ob["column6"] # Cliente
    self.worksheet.cell(column=7, row=row).value = ob["column7"] # Cobrança Vinculada / Defeito
    self.worksheet.cell(column=8, row=row).value = ob["column8"] # Distribuida
    self.worksheet.cell(column=9, row=row).value = ob["column9"] # Cidade
    self.worksheet.cell(column=10, row=row).value = ob["column10"] # Bairro

  def add_into_sheet(self, obj_list):
    for i, ob in enumerate(obj_list):
      self.add_ob(ob, i + 2)
  
  def save(self, path_file):
    self.workbook.save(path_file)


data = date.today() + timedelta(days=1)

dt_month = data.strftime('%m.%Y')
dt_file = data.strftime('%d.%m.%Y')
tomorrow = data.strftime('%d/%m/%Y')

path_xlsx = '/home/douglas/dev/py/intranet/app/static/xlsx/model.xlsx' # Arquivo .xlsx
path_folder = f'/home/douglas/Documentos/worksheet/{dt_month}'         # Onde vai salvar
path_file = f'{path_folder}/{dt_file}.xlsx'

if path.isdir(path_folder):
  pass
else:
  makedirs(path_folder)

connect = databaseConnection("dbname='mkData3.0' user='cliente_r' host='177.184.72.6' password='Cl13nt_R'")
createWorksheet = workSheet(path_xlsx)

database = connect.Consult("""SELECT os.codos, df.descricao_defeito, tp.descricao, os.data_abertura, cl.nome_razaosocial, cd.cidade, ba.bairro, lo.logradouro
  FROM mk_os os
  FULL OUTER JOIN mk_os_tipo tp ON os.tipo_os = tp.codostipo
  JOIN mk_pessoas cl ON os.cliente = cl.codpessoa
  JOIN mk_os_defeitos df ON os.defeito_associado = df.coddefeito
  JOIN mk_cidades cd ON os.cd_cidade = cd.codcidade
  JOIN mk_bairros ba ON os.cd_bairro = ba.codbairro
  JOIN mk_logradouros lo ON os.cd_logradouro = lo.codlogradouro
  WHERE status='1' AND tipo_os in ('4','15','18') AND fechamento_tecnico='N' ORDER BY cd.cidade asc""")

obj_list = createWorksheet.dumpData(database, tomorrow)
createWorksheet.add_into_sheet(obj_list)
createWorksheet.save(path_file)

webbrowser.open(path.abspath(path_file))
connect.close()
